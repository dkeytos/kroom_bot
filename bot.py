import logging
import time
from telegram.ext import (
    ApplicationBuilder, CommandHandler, MessageHandler, filters, ConversationHandler, CallbackQueryHandler
)
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import ContextTypes
import re  # For email validation
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from dotenv import load_dotenv
load_dotenv()

TOKEN = os.getenv("BOT_TOKEN")
OWNER_CHAT_ID = int(os.getenv("OWNER_CHAT_ID"))

def is_valid_email(email: str) -> bool:
    # Simple regex for email validation
    pattern = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
    return re.match(pattern, email) is not None

# Set up logging so you can see what’s happening
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
# Suppress httpx debug messages:
logging.getLogger("httpx").setLevel(logging.WARNING)

# Conversation state constants
WAITING_FOR_USERNAME = 0
WAITING_FOR_PHOTO = 1
CHOOSING_OPTION = 2
WAITING_FOR_EMAIL = 3  # New state for registered email input

# Add mapping for language flags
lang_flags = {
    "eng": "🇬🇧",
    "ita": "🇮🇹",
    "spa": "🇪🇸"
}

# Add language-specific messages and helper function
language_msgs = {
    "eng": {
        "ask_username": "Please enter your Keytos username 😊:",
        "unset_username": "⚠️ Your Telegram username is not set. Please update your Telegram profile and send 'OK'.",
        "ask_photo": "Hey there! 😊 Please send a photo as proof of deposit, as the ones shown above. 📸 Only a screenshot is accepted!\n\nPlease note, only deposit >300$ grant access to KeyRoom.",
        "invalid_photo": "Oops! 😕 Kindly note, only a photo can be used as proof of deposit. 📸 Please send a picture. 👍",
        "invalid_photo_reset": "Oops! 😕 Only a photo works. Please send a picture.\n\nOr press {reset_button} to start over!",
        "success": "Awesome! 😊 Our support team will contact you shortly to get you into KeyRoom! 🚀",
        "choose_option": "Please choose an option below: 👇",
        "deposit_proof_button": "Deposit Proof 📸",
        "already_registered_button": "Already Registered ✅",
        "reset_button": "Reset 🔄",
        "ask_email": "Great! Now, please enter your AXI registered email address 📧",
        "invalid_email": "Hmm... That doesn't look like a valid email address 😕. Please send a valid email address 📧"
    },
    "ita": {
        "ask_username": "Per favore, inserisci il tuo username di Keytos 😊:",
        "unset_username": "⚠️ Il tuo username Telegram non è impostato. Aggiorna il tuo profilo Telegram e invia 'OK'.",
        "ask_photo": "Ciao! 😊 Per favore, invia una foto come prova del deposito, come quelle mostrate qui sopra. 📸 Solo schreenshots sono accettati!\n\nRicorda, solo i depositi >300$ danno accesso a KeyRoom.",
        "invalid_photo": "Ops! 😕 Nota bene, solo una foto può essere usata come prova del deposito. 📸 Per favore, invia una foto. 👍",
        "invalid_photo_reset": "Ops! 😕 Solo una foto funziona. Invia una foto.\n\nO premi {reset_button} per ricominciare!",
        "success": "Fantastico! 😊 Il nostro team di supporto ti contatterà a breve per farti entrare in KeyRoom! 🚀",
        "choose_option": "Per favore, scegli un'opzione qui sotto: 👇",
        "deposit_proof_button": "Prova di deposito 📸",
        "already_registered_button": "Già registrato ✅",
        "reset_button": "Ricomincia 🔄",
        "ask_email": "Perfetto! Ora, inserisci l'indirizzo email con cui ti sei registrato su AXI 📧",
        "invalid_email": "Ops... L'indirizzo email non sembra valido 😕. Invia un indirizzo email valido 📧"
    },
    "spa": {
        "ask_username": "Por favor, ingresa tu nombre de usuario de Keytos 😊:",
        "unset_username": "⚠️ Tu nombre de usuario de Telegram no está configurado. Actualiza tu perfil de Telegram y enviar 'OK'.",
        "ask_photo": "¡Hola! 😊 Por favor, envía una foto como prueba del depósito, como las que se muestran arriba. 📸 ¡Sólo se aceptan capturas de pantalla.!\n\nRecuerde que sólo los depósitos >300$ dan acceso a KeyRoom.",
        "invalid_photo": "Uy! 😕 Tenga en cuenta que solo se puede usar una foto como prueba del depósito. 📸 Por favor, envíe una imagen. 👍",
        "invalid_photo_reset": "Uy! 😕 Solo se acepta una imagen. Envíe una imagen.\n\nO presione {reset_button} para reiniciar!",
        "success": "¡Genial! 😊 Nuestro equipo de soporte se pondrá en contacto con usted en breve para que pueda ingresar a KeyRoom! 🚀",
        "choose_option": "Por favor, elija una opción a continuación: 👇",
        "deposit_proof_button": "Comprobante de depósito 📸",
        "already_registered_button": "Ya registrado ✅",
        "reset_button": "Reiniciar 🔄",
        "ask_email": "¡Perfecto! Ahora, ingrese el correo electrónico con el que se registró en AXI📧",
        "invalid_email": "¡Uy! 😕 El correo electrónico ingresado no es válido. Por favor, ingrese un correo electrónico válido 📧"
    }
}

def get_language(context: ContextTypes.DEFAULT_TYPE):
    reg_param = context.user_data.get("reg_param", "")
    if "_" in reg_param:
        return reg_param.split("_")[0]
    return "eng"

# Updated helper function to save contact with new columns.
def save_contact_to_excel(lead: dict):
    # Determine the file name based on today's date.
    today = datetime.now().strftime("%Y-%m-%d")
    # Save in a leads subfolder.
    folder = os.path.join(os.getcwd(), "leads")
    if not os.path.exists(folder):
        os.makedirs(folder)
    file_path = os.path.join(folder, f"{today}.xlsx")
    
    # New headers: Date, Telegram Username, Keytos Username, User ID, Flow, Email, Language, Proof Picture
    headers = ["Date", "Telegram Username", "Keytos Username", "Flow", "Email", "Language"]
    
    if os.path.exists(file_path):
        wb = load_workbook(file_path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
    
    # Prepare the row data.
    row = [
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        lead.get("telegram_username", ""),
        lead.get("keytos_username", ""),
        lead.get("flow", ""),
        lead.get("email", ""),
        lead.get("language", "")
    ]
    ws.append(row)
    wb.save(file_path)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    args = update.message.text.split()
    param = args[1] if len(args) > 1 else "Not provided"
    if param == "Not provided":
        # Delete /start message and show the main menu as usual.
        try:
            await context.bot.delete_message(chat_id=update.effective_chat.id, message_id=update.message.message_id)
        except Exception as e:
            logger.error("Failed to delete /start command message: %s", e)
        prev_menu_msg_id = context.user_data.get("menu_msg_id")
        if prev_menu_msg_id:
            try:
                await context.bot.delete_message(chat_id=update.effective_chat.id, message_id=prev_menu_msg_id)
            except Exception as e:
                logger.error("Failed to delete previous menu: %s", e)
        keyboard = [
            [InlineKeyboardButton("🇬🇧", callback_data="eng"),
             InlineKeyboardButton("🇮🇹", callback_data="ita"),
             InlineKeyboardButton("🇪🇸", callback_data="spa")],
            [InlineKeyboardButton("US Residents", callback_data="us")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        new_menu = await update.message.reply_text(language_msgs["eng"]["choose_option"], reply_markup=reply_markup)
        context.user_data["menu_msg_id"] = new_menu.message_id
        return CHOOSING_OPTION
    else:
        # If a parameter is provided, check for us_resident first.
        if param == "us_resident":
            context.user_data["reg_param"] = param
            context.user_data["lang"] = "eng"
            context.user_data["flow"] = "us"
            await update.message.reply_text(language_msgs["eng"]["ask_username"])
            return WAITING_FOR_USERNAME
        elif param.endswith("_deposit"):
            lang = param.split("_")[0]
            context.user_data["lang"] = lang
            # Send sample pictures before asking for the photo.
            ts = int(time.time())
            pc_sample_url = f"https://keyroom-images-bucket.s3.eu-central-1.amazonaws.com/{lang}_pc.png?v={ts}"
            mobile_sample_url = f"https://keyroom-images-bucket.s3.eu-central-1.amazonaws.com/{lang}_mobile.png?v={ts}"
            await update.message.reply_photo(photo=pc_sample_url, caption="PC Screenshot")
            await update.message.reply_photo(photo=mobile_sample_url, caption="Mobile Screenshot")
            # Prompt for deposit proof in the chosen language.
            await update.message.reply_text(language_msgs[lang]["ask_photo"])
            context.user_data["flow"] = "deposit"
            return WAITING_FOR_PHOTO
        elif param.endswith("_register"):
            lang = param.split("_")[0]
            context.user_data["lang"] = lang
            await update.message.reply_text(language_msgs[lang]["ask_email"])
            context.user_data["flow"] = "register"
            return WAITING_FOR_EMAIL
        else:
            # Fallback: show the main menu.
            return await send_start_menu(update, context)

async def send_start_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keyboard = [
        [InlineKeyboardButton("🇬🇧", callback_data="eng"),
         InlineKeyboardButton("🇮🇹", callback_data="ita"),
         InlineKeyboardButton("🇪🇸", callback_data="spa")],
        [InlineKeyboardButton("US Residents", callback_data="us")]
    ]
    reply_markup = InlineKeyboardMarkup(keyboard)
    # If the reset came from a callback, use edit_message_text
    if update.callback_query:
        await update.callback_query.edit_message_text(language_msgs["eng"]["choose_option"], reply_markup=reply_markup)
    else:
        await update.message.reply_text(language_msgs["eng"]["choose_option"], reply_markup=reply_markup)
    return CHOOSING_OPTION

async def choice_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data in ["eng", "ita", "spa"]:
        context.user_data["lang"] = query.data  
        context.user_data["reg_param"] = query.data + "_"  # tentative action
        keyboard = [
            [InlineKeyboardButton(language_msgs[query.data]["deposit_proof_button"], callback_data="deposit_proof"),
             InlineKeyboardButton(language_msgs[query.data]["already_registered_button"], callback_data="already_registered")],
            [InlineKeyboardButton(language_msgs[query.data]["reset_button"], callback_data="reset")]
        ]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await query.edit_message_text(language_msgs[query.data]["choose_option"], reply_markup=reply_markup)
        return CHOOSING_OPTION
    elif query.data == "us":
        context.user_data["reg_param"] = "us_resident"
        context.user_data["lang"] = "eng"
        # Directly ask for the username for US residents.
        context.user_data["flow"] = "us"
        await query.edit_message_text(language_msgs["eng"]["ask_username"])
        return WAITING_FOR_USERNAME
    elif query.data == "deposit_proof":
        lang = context.user_data.get("lang", "eng")
        context.user_data["reg_param"] = lang + "_deposit"
        ts = int(time.time())
        pc_sample_url = f"https://keyroom-images-bucket.s3.eu-central-1.amazonaws.com/{lang}_pc.png?v={ts}"
        mobile_sample_url = f"https://keyroom-images-bucket.s3.eu-central-1.amazonaws.com/{lang}_mobile.png?v={ts}"
        await context.bot.send_photo(chat_id=update.effective_chat.id, photo=pc_sample_url, caption="PC Screenshot")
        await context.bot.send_photo(chat_id=update.effective_chat.id, photo=mobile_sample_url, caption="Mobile Screenshot")
        await query.edit_message_text(language_msgs[lang]["ask_photo"])
        return WAITING_FOR_PHOTO
    elif query.data == "already_registered":
        lang = context.user_data.get("lang", "eng")
        context.user_data["reg_param"] = lang + "_register"
        await query.edit_message_text(language_msgs[lang]["ask_email"])
        return WAITING_FOR_EMAIL
    elif query.data == "reset":
        return await send_start_menu(update, context)

# In keytos_username_handler, update the calls to save_contact_to_excel:

async def keytos_username_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Check if the Telegram username is set.
    if not update.message.from_user.username:
        lang = context.user_data.get("lang", "eng")
        await update.message.reply_text(language_msgs[lang]["unset_username"])
        return WAITING_FOR_USERNAME

    keytos_username = update.message.text.strip()
    context.user_data["keytos_username"] = keytos_username
    flow = context.user_data.get("flow")
    user = update.message.from_user
    username = f"@{user.username}"  # already set
    user_id = user.id
    lang = context.user_data.get("lang", "eng")
    
    if flow == "us":
        forward_msg = (
            "New Contact (US Resident):\n"
            "Username: " + username + "\n" +
            "Keytos Username: " + keytos_username
        )
        await context.bot.send_message(chat_id=OWNER_CHAT_ID, text=forward_msg)
        await update.message.reply_text(language_msgs[lang]["success"])
        logger.info("Forwarded contact info (US Resident): %s", forward_msg)
        save_contact_to_excel({
            "telegram_username": username,
            "keytos_username": keytos_username,
            "flow": flow,
            "language": lang
        })
        return ConversationHandler.END

    elif flow == "deposit":
        forward_msg = (
            "New Deposit Proof:\n"
            "Username: " + username + "\n" +
            "Keytos Username: " + keytos_username + "\n"
        )
        await context.bot.send_photo(chat_id=OWNER_CHAT_ID, photo=context.user_data["deposit_photo"], caption=forward_msg)
        await update.message.reply_text(language_msgs[lang]["success"])
        logger.info("Forwarded deposit proof from user %s", username)
        save_contact_to_excel({
            "telegram_username": username,
            "keytos_username": keytos_username,
            "flow": flow,
            "language": lang
        })
        return ConversationHandler.END

    elif flow == "register":
        email = context.user_data.get("email", "Not provided")
        language_line = ""
        if lang in lang_flags:
            language_line = " " + lang_flags[lang]
        forward_msg = (
            "New Contact (Already Registered):\n"
            "Username: " + username + "\n" +
            "Keytos Username: " + keytos_username + "\n" +
            "Email: " + email + language_line
        )
        await context.bot.send_message(chat_id=OWNER_CHAT_ID, text=forward_msg)
        await update.message.reply_text(language_msgs[lang]["success"])
        logger.info("Forwarded contact info (Register): %s", forward_msg)
        save_contact_to_excel({
            "telegram_username": username,
            "keytos_username": keytos_username,
            "flow": flow,
            "email": email,
            "language": lang
        })
        return ConversationHandler.END

    else:
        forward_msg = (
            "New Contact:\n"
            "Username: " + username + "\n" +
            "User ID: " + str(user_id) + "\n" +
            "Keytos Username: " + keytos_username
        )
        await context.bot.send_message(chat_id=OWNER_CHAT_ID, text=forward_msg)
        await update.message.reply_text(language_msgs[lang]["success"])
        logger.info("Forwarded contact info: %s", forward_msg)
        save_contact_to_excel({
            "telegram_username": username,
            "keytos_username": keytos_username,
            "flow": flow if flow else "contact",
            "email": "",
            "language": lang
        })
        return ConversationHandler.END

async def deposit_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not update.message.photo:
        lang = get_language(context)
        await update.message.reply_text(language_msgs[lang]["invalid_photo"])
        return WAITING_FOR_PHOTO
    # Select the highest-resolution photo available.
    photo = update.message.photo[-1]
    user = update.message.from_user
    username = f"@{user.username}" if user.username else "Not set"
    caption = "New Deposit Proof:\nUsername: " + username
    lang = context.user_data.get("lang", "eng")
    if lang in lang_flags:
        caption += "\nLanguage: " + lang_flags[lang]
    # Store the photo info and set the deposit flow.
    context.user_data["deposit_photo"] = photo.file_id
    context.user_data["deposit_caption"] = caption
    context.user_data["flow"] = "deposit"
    # Ask for the Keytos username in the chosen language.
    await update.message.reply_text(language_msgs[lang]["ask_username"])
    return WAITING_FOR_USERNAME

async def deposit_invalid(update: Update, context: ContextTypes.DEFAULT_TYPE):
    lang = get_language(context)
    keyboard = [[InlineKeyboardButton(language_msgs[lang]["reset_button"], callback_data="reset")]]
    reply_markup = InlineKeyboardMarkup(keyboard)
    error_msg = language_msgs[lang]["invalid_photo_reset"].format(reset_button=language_msgs[lang]["reset_button"])
    await update.message.reply_text(error_msg, reply_markup=reply_markup)
    return WAITING_FOR_PHOTO

async def registered_email_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    email_input = update.message.text.strip()
    reg_param = context.user_data.get("reg_param", "")
    lang = reg_param.split("_")[0] if "_" in reg_param else "eng"
    if not is_valid_email(email_input):
        keyboard = [[InlineKeyboardButton(language_msgs[lang]["reset_button"], callback_data="reset")]]
        reply_markup = InlineKeyboardMarkup(keyboard)
        await update.message.reply_text(language_msgs[lang]["invalid_email"], reply_markup=reply_markup)
        return WAITING_FOR_EMAIL
    # Store the email and set the flow.
    context.user_data["email"] = email_input
    context.user_data["flow"] = "register"
    await update.message.reply_text(language_msgs[lang]["ask_username"])
    return WAITING_FOR_USERNAME

async def echo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # Instruct user to use /start instead of sending arbitrary texts.
    await update.message.reply_text("Please type /start to begin the conversation.")

# Modify reset_command to call send_start_menu
async def reset_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    logger.info("Reset command received.")
    return await send_start_menu(update, context)

async def main():
    application = ApplicationBuilder().token(TOKEN).build()
    
    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            WAITING_FOR_USERNAME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, keytos_username_handler)
            ],
            CHOOSING_OPTION: [CallbackQueryHandler(choice_handler)],
            WAITING_FOR_PHOTO: [
                CallbackQueryHandler(reset_command, pattern='^reset$'),
                MessageHandler(filters.PHOTO, deposit_photo),
                MessageHandler(~filters.PHOTO, deposit_invalid)
            ],
            WAITING_FOR_EMAIL: [
                CallbackQueryHandler(reset_command, pattern='^reset$'),  # New handler for reset in WAITING_FOR_EMAIL
                MessageHandler(filters.TEXT & ~filters.COMMAND, registered_email_handler)
            ]
        },
        fallbacks=[CommandHandler("reset", reset_command)],
        allow_reentry=True  # Allow /start to be processed even if conversation is active.
    )
    application.add_handler(conv_handler)
    # Replace the echo handler to force /start usage.
    application.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, echo))
    
    await application.run_polling()

if __name__ == '__main__':
    import asyncio
    try:
        import nest_asyncio
        nest_asyncio.apply()
    except ImportError:
        pass
    try:
        asyncio.run(main())
    except RuntimeError as e:
        if "Cannot close a running event loop" in str(e):
            logger.warning("Event loop error suppressed: %s", e)
        else:
            raise
